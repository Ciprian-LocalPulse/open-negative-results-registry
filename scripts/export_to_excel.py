#!/usr/bin/env python3
"""
export_to_excel.py

Exports the entire negative-result database (all JSON entries under data/)
into a single, well-formatted Excel workbook with one sheet per domain
plus a combined "All Data" sheet.

Usage:
    python export_to_excel.py
    python export_to_excel.py --output DarkData_Export.xlsx

Requires:
    pip install openpyxl
"""

import argparse
import json
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependency. Install with: pip install openpyxl --break-system-packages")
    raise SystemExit(1)

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"
EXCLUDED_DIRS = {"templates"}

COLUMNS = [
    "experiment_id", "domain", "target_disease", "intervention_type",
    "intervention_name", "dosage", "hypothesis", "outcome",
    "methodology_summary", "researcher_orcid", "institution_type",
    "date_concluded", "source_type", "source_url", "license",
]

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def flatten(entry: dict) -> list:
    intervention = entry.get("tested_intervention", {})
    return [
        entry.get("experiment_id", ""),
        entry.get("domain", ""),
        entry.get("target_disease", ""),
        intervention.get("type", ""),
        intervention.get("name", ""),
        intervention.get("dosage", ""),
        entry.get("hypothesis", ""),
        entry.get("outcome", ""),
        entry.get("methodology_summary", ""),
        entry.get("researcher_orcid", ""),
        entry.get("institution_type", ""),
        entry.get("date_concluded", ""),
        entry.get("source_type", ""),
        entry.get("source_url", ""),
        entry.get("license", ""),
    ]


def style_sheet(ws):
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 22
    ws.freeze_panes = "A2"


def load_entries_by_domain():
    by_domain = {}
    for domain_dir in DATA_DIR.iterdir():
        if not domain_dir.is_dir() or domain_dir.name in EXCLUDED_DIRS:
            continue
        entries = []
        for json_file in sorted(domain_dir.glob("*.json")):
            try:
                with open(json_file, "r", encoding="utf-8") as f:
                    entries.append(json.load(f))
            except json.JSONDecodeError:
                print(f"Skipping invalid JSON: {json_file}")
        if entries:
            by_domain[domain_dir.name] = entries
    return by_domain


def main():
    parser = argparse.ArgumentParser(description="Export the negative-result database to Excel.")
    parser.add_argument("--output", default="DarkData_Export.xlsx", help="Output .xlsx path")
    args = parser.parse_args()

    by_domain = load_entries_by_domain()
    if not by_domain:
        print("No data found yet. Add submissions under data/<domain>/ first.")
        return

    wb = Workbook()
    all_sheet = wb.active
    all_sheet.title = "All Data"
    style_sheet(all_sheet)

    row_cursor = 2
    for domain, entries in sorted(by_domain.items()):
        sheet_name = domain[:31]  # Excel sheet name limit
        ws = wb.create_sheet(title=sheet_name)
        style_sheet(ws)
        for r, entry in enumerate(entries, start=2):
            row_data = flatten(entry)
            for c, value in enumerate(row_data, start=1):
                ws.cell(row=r, column=c, value=value)
            for c, value in enumerate(row_data, start=1):
                all_sheet.cell(row=row_cursor, column=c, value=value)
            row_cursor += 1

    wb.save(args.output)
    total = sum(len(v) for v in by_domain.values())
    print(f"Exported {total} entries across {len(by_domain)} domains to {args.output}")


if __name__ == "__main__":
    main()
