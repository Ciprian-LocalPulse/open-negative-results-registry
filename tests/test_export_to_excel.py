#!/usr/bin/env python3
"""
test_export_to_excel.py

Covers scripts/export_to_excel.py, which had zero test coverage.
"""
import importlib
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / "scripts"))

import export_to_excel as ete  # noqa: E402

SAMPLE_ENTRY = {
    "experiment_id": "TEST-0001",
    "domain": "Oncology",
    "target_disease": "Test Disease",
    "tested_intervention": {"type": "Drug", "name": "Test Compound", "dosage": "50mg"},
    "hypothesis": "Reduces tumor growth.",
    "outcome": "No significant effect.",
    "methodology_summary": "RCT, n=100.",
    "researcher_orcid": "0000-0002-1825-0097",
    "institution_type": "University Research Lab",
    "date_concluded": "2025-01-01",
    "source_type": "original_submission",
    "source_url": "https://example.org",
    "license": "CC0-1.0",
}


def test_flatten_maps_all_columns_in_order():
    row = ete.flatten(SAMPLE_ENTRY)
    assert row[0] == "TEST-0001"
    assert row[1] == "Oncology"
    assert row[4] == "Test Compound"  # intervention_name
    assert row[5] == "50mg"  # dosage
    assert len(row) == len(ete.COLUMNS)


def test_flatten_handles_missing_intervention_gracefully():
    entry = dict(SAMPLE_ENTRY)
    del entry["tested_intervention"]
    row = ete.flatten(entry)
    assert row[4] == ""  # intervention_name defaults empty, no KeyError


def test_load_entries_by_domain_skips_invalid_json(tmp_path, monkeypatch, capsys):
    domain_dir = tmp_path / "oncology"
    domain_dir.mkdir()
    (domain_dir / "good.json").write_text(json.dumps(SAMPLE_ENTRY), encoding="utf-8")
    (domain_dir / "broken.json").write_text("{not json", encoding="utf-8")
    monkeypatch.setattr(ete, "DATA_DIR", tmp_path)

    by_domain = ete.load_entries_by_domain()

    assert len(by_domain["oncology"]) == 1
    assert "Skipping invalid JSON" in capsys.readouterr().out


def test_load_entries_by_domain_excludes_templates(tmp_path, monkeypatch):
    (tmp_path / "templates").mkdir()
    (tmp_path / "templates" / "t.json").write_text(json.dumps(SAMPLE_ENTRY), encoding="utf-8")
    monkeypatch.setattr(ete, "DATA_DIR", tmp_path)

    by_domain = ete.load_entries_by_domain()

    assert "templates" not in by_domain


def test_main_writes_workbook_with_all_data_and_domain_sheets(tmp_path, monkeypatch):
    domain_dir = tmp_path / "oncology"
    domain_dir.mkdir()
    (domain_dir / "e1.json").write_text(json.dumps(SAMPLE_ENTRY), encoding="utf-8")
    monkeypatch.setattr(ete, "DATA_DIR", tmp_path)

    out_path = tmp_path / "out.xlsx"
    monkeypatch.setattr(sys, "argv", ["export_to_excel.py", "--output", str(out_path)])
    ete.main()

    assert out_path.exists()
    wb = load_workbook(out_path)
    assert "All Data" in wb.sheetnames
    assert "oncology" in wb.sheetnames
    assert wb["All Data"].cell(row=2, column=1).value == "TEST-0001"


def test_main_with_no_data_does_not_crash(tmp_path, monkeypatch, capsys):
    monkeypatch.setattr(ete, "DATA_DIR", tmp_path)  # empty dir, no domain subfolders
    monkeypatch.setattr(sys, "argv", ["export_to_excel.py"])
    ete.main()
    assert "No data found" in capsys.readouterr().out
